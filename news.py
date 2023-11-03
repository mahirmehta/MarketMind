import requests
import openpyxl
from openpyxl import load_workbook

# Replace with your News API key
NEWS_API_KEY = ' '

# List of bank stock symbols you want to search for (both Indian and non-Indian banks)
bank_symbols = ['HDFCBANK', 'ICICIBANK', 'AXISBANK', 'KOTAKBANK', 'PNB', 'AUBANK', 'BANKBARODA', 'INDUSINDBK', 'FEDERALBNK', 'SBIN', 'BANDHANBNK','IDFCFIRSTB']

# Load an existing Excel workbook or create a new one if it doesn't exist
try:
    workbook = load_workbook('bank_stock_news.xlsx')
    worksheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Bank Stock News'
    worksheet['A1'] = 'Symbol'
    worksheet['B1'] = 'Title'
    worksheet['C1'] = 'Description'
    worksheet['D1'] = 'URL'
    worksheet['E1'] = 'Date and Time'

# Find the last row in the Excel sheet to append new data
row_num = worksheet.max_row + 1

# Loop through each bank stock symbol and fetch news
for stock_symbol in bank_symbols:
    # Define the News API URL for the current stock symbol
    news_api_url = f'https://newsapi.org/v2/everything?q={stock_symbol}&apiKey={NEWS_API_KEY}'

    # Make the API request
    response = requests.get(news_api_url)

    if response.status_code == 200:
        news_data = response.json()
        articles = news_data.get('articles', [])

        for article in articles:
            title = article.get('title', '')
            description = article.get('description', '')
            url = article.get('url', '')
            published_at = article.get('publishedAt', '')

            # Write data to the Excel file
            worksheet[f'A{row_num}'] = stock_symbol
            worksheet[f'B{row_num}'] = title
            worksheet[f'C{row_num}'] = description
            worksheet[f'D{row_num}'] = url
            worksheet[f'E{row_num}'] = published_at

            row_num += 1
    else:
        print(f'Failed to fetch news for {stock_symbol}')

# Save the updated Excel file
workbook.save('bank_stock_news.xlsx')
print('Bank stock news appended to bank_stock_news.xlsx')